from sqlalchemy.orm import Session
from datetime import datetime
from typing import List, Dict, Any
from io import BytesIO
import json
import crud
import models


class ArtifactGenerator:
    """成果物生成エンジン"""
    
    def __init__(self, db: Session, project_id: int):
        self.db = db
        self.project_id = project_id
        self._load_data()
    
    def _load_data(self):
        """必要なデータを読み込む"""
        self.project = crud.get_project(self.db, self.project_id)
        self.decisions = crud.get_decisions(self.db, self.project_id)
        self.backlog_items = crud.get_backlog_items(self.db, self.project_id)
        self.answers = crud.get_answers(self.db, self.project_id)
        
        # 設定項目マスタ
        all_config_items = crud.get_config_items(self.db)
        self.config_items = {item.id: item for item in all_config_items}
        
        # 回答済み設定項目
        self.answered_config_ids = set(answer.config_item_id for answer in self.answers)
        
        # 回答をconfig_item_id別に整理
        self.answers_by_config = {}
        for answer in self.answers:
            if answer.config_item_id not in self.answers_by_config:
                self.answers_by_config[answer.config_item_id] = []
            self.answers_by_config[answer.config_item_id].append(answer)
    
    def _count_tbd(self, content: str) -> int:
        """TBD（未決定）の数をカウント"""
        return content.count('TBD') + content.count('未決定')
    
    def generate_decision_log(self) -> str:
        """
        Decision Log（決定事項ログ）を生成
        
        全ての決定事項を時系列で記録
        """
        lines = [
            "# Decision Log（決定事項ログ）",
            "",
            f"**プロジェクト**: {self.project.name}",
            f"**生成日時**: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC",
            "",
            "---",
            ""
        ]
        
        if not self.decisions:
            lines.append("## まだ決定事項がありません")
            lines.append("")
            lines.append("ウィザードで質問に回答すると、ここに決定事項が記録されます。")
        else:
            lines.append("## 決定事項一覧")
            lines.append("")
            
            for i, decision in enumerate(self.decisions, 1):
                config_item = self.config_items.get(decision.config_item_id)
                
                lines.append(f"### {i}. {decision.title}")
                lines.append("")
                lines.append(f"- **設定項目ID**: {decision.config_item_id}")
                if config_item:
                    lines.append(f"- **優先度**: {config_item.priority}")
                lines.append(f"- **決定日時**: {decision.created_at.strftime('%Y-%m-%d %H:%M:%S')}")
                lines.append(f"- **ステータス**: {decision.status}")
                lines.append("")
                
                if decision.rationale:
                    lines.append("**決定内容**:")
                    lines.append("")
                    lines.append(decision.rationale)
                    lines.append("")
                
                if decision.impact:
                    lines.append("**影響範囲**:")
                    lines.append("")
                    lines.append(decision.impact)
                    lines.append("")
                
                lines.append("---")
                lines.append("")
        
        return "\n".join(lines)
    
    def generate_config_workbook(self) -> str:
        """
        Config Workbook（設定作業一覧）を生成
        
        必要な設定項目を一覧表示（ID、タイトル、ステータス、依存関係）
        """
        lines = [
            "# Config Workbook（設定作業一覧）",
            "",
            f"**プロジェクト**: {self.project.name}",
            f"**生成日時**: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC",
            "",
            "---",
            ""
        ]
        
        # サマリー
        total = len(self.backlog_items)
        done = sum(1 for item in self.backlog_items if item.status == models.BacklogStatus.DONE)
        ready = sum(1 for item in self.backlog_items if item.status == models.BacklogStatus.READY)
        blocked = sum(1 for item in self.backlog_items if item.status == models.BacklogStatus.BLOCKED)
        
        lines.append("## サマリー")
        lines.append("")
        lines.append(f"- **全設定項目数**: {total}")
        lines.append(f"- **完了**: {done} ({round(done/total*100, 1) if total > 0 else 0}%)")
        lines.append(f"- **対応可能**: {ready}")
        lines.append(f"- **ブロック中**: {blocked}")
        lines.append("")
        lines.append("---")
        lines.append("")
        
        # 優先度別に整理
        items_by_priority = {}
        for item in self.backlog_items:
            config_item = self.config_items.get(item.config_item_id)
            if config_item:
                priority = config_item.priority or 'UNKNOWN'
                if priority not in items_by_priority:
                    items_by_priority[priority] = []
                items_by_priority[priority].append((item, config_item))
        
        # 優先度順に出力
        for priority in ['P0', 'P1', 'P2', 'P3']:
            if priority not in items_by_priority:
                continue
            
            lines.append(f"## 優先度: {priority}")
            lines.append("")
            lines.append("| ID | タイトル | ステータス | 依存関係 | 設定値 |")
            lines.append("|---|---|---|---|---|")
            
            for backlog_item, config_item in items_by_priority[priority]:
                status_emoji = {
                    models.BacklogStatus.DONE: "✅",
                    models.BacklogStatus.READY: "🟢",
                    models.BacklogStatus.BLOCKED: "🔴",
                    models.BacklogStatus.PENDING: "⚪"
                }.get(backlog_item.status, "❓")
                
                # 依存関係
                depends = config_item.depends_on or []
                depends_str = ", ".join(depends) if depends else "-"
                
                # 設定値
                if backlog_item.answered:
                    answers = self.answers_by_config.get(config_item.id, [])
                    value_parts = []
                    for ans in answers:
                        val = ans.value
                        if isinstance(val, list):
                            val = ", ".join(str(v) for v in val)
                        value_parts.append(f"{ans.input_name}={val}")
                    value_str = "; ".join(value_parts) if value_parts else "設定済み"
                else:
                    value_str = "**TBD（未決定）**"
                
                lines.append(
                    f"| {config_item.id} | {config_item.title} | "
                    f"{status_emoji} {backlog_item.status.value} | {depends_str} | {value_str} |"
                )
            
            lines.append("")
        
        return "\n".join(lines)
    
    # 設定項目IDごとの固有テスト観点マッピング
    _TEST_PERSPECTIVES = {
        'FI-CORE-001': [
            "会計年度が正しく設定され、期間が12ヶ月に分割されているか確認",
            "年度を跨ぐ転記（3月末→4月初）が正しく処理されるか確認",
            "特別期間（調整期間）の利用可否を確認",
            "年度末の繰越処理が正常に動作するか確認",
        ],
        'FI-CORE-002': [
            "会社コードでの転記が正しい通貨で記録されるか確認",
            "外貨取引時に為替レートが適用されるか確認",
            "国設定に基づく税制が正しくリンクされるか確認",
            "複数会社コード間の会社間取引が正しく処理されるか確認",
        ],
        'FI-CORE-003': [
            "締め後に通常ユーザーが転記できないことを確認",
            "期間オープン/クローズの切替が正しく動作するか確認",
            "例外許可設定が正しく機能するか確認",
            "締め前の未転記伝票が警告されるか確認",
        ],
        'FI-CORE-004': [
            "伝票タイプごとに採番が正しく連番で付与されるか確認",
            "年度切替時に採番がリセットされるか確認",
            "採番キーの組合せ（会社コード×年度×タイプ）が正しく動作するか確認",
            "採番の重複が発生しないことを確認（同時転記テスト）",
        ],
        'FI-CORE-005': [
            "BS/PL区分が正しく設定されているか確認",
            "オープンアイテム管理対象の勘定で未消込明細が表示されるか確認",
            "統制勘定への直接転記が禁止されているか確認",
            "勘定科目グループの分類が正しいか確認",
        ],
        'FI-APAR-001': [
            "BPマスタの登録・変更・照会が正しく動作するか確認",
            "得意先/仕入先ロールが正しく設定されるか確認",
            "BP番号の採番ルールが正しく動作するか確認",
            "重複BPのチェック機能が動作するか確認",
        ],
        'FI-APAR-002': [
            "統制勘定とサブレジャの残高が一致するか確認",
            "統制勘定変更時の影響が正しく反映されるか確認",
            "BPグループ別の統制勘定割当が正しいか確認",
            "統制勘定への直接転記が禁止されているか確認",
        ],
        'FI-APAR-003': [
            "支払条件に基づく支払期限が正しく計算されるか確認",
            "消込処理（全額・部分）が正しく動作するか確認",
            "支払方法（振込・手形等）の切替が正しく動作するか確認",
            "締め日をまたぐ取引の支払期限計算を確認",
        ],
        'FI-TAX-001': [
            "標準税率（10%）と軽減税率（8%）が正しく計算されるか確認",
            "内税/外税の切替が正しく動作するか確認",
            "税計算の端数処理（四捨五入/切り捨て/切り上げ）が正しいか確認",
            "非課税・免税取引が正しく処理されるか確認",
            "インボイス制度（適格請求書）への対応を確認",
        ],
        'FI-DIFF-001': [
            "許容差額内の差額が自動消込されるか確認",
            "許容差額を超える差額が手動消込に回されるか確認",
            "差額勘定への自動転記が正しく行われるか確認",
            "通貨端数処理が設定通りに動作するか確認",
        ],
        'FI-DIFF-002': [
            "部分入金時の残余アイテムが正しく生成されるか確認",
            "過払い時の処理（クレジットメモ/返金）が正しく動作するか確認",
            "不足払い時の残余処理方針が設定通りに動作するか確認",
            "複数明細の一括消込が正しく動作するか確認",
        ],
        'FI-CLOSE-001': [
            "例外転記に適切な承認が要求されるか確認",
            "監査ログに例外操作が記録されるか確認",
            "承認なしの例外転記が拒否されるか確認",
            "監査ログの保持期間設定が正しいか確認",
            "期末決算時の例外権限が正しく管理されるか確認",
        ],
        'FI-RPT-001': [
            "試算表の貸借が一致するか確認",
            "未消込一覧に正しい明細が表示されるか確認",
            "エイジング分析の期間区分が正しいか確認",
            "レポート出力形式（画面/PDF/Excel/CSV）が正しく動作するか確認",
            "月次/週次のレポート自動生成が正しくスケジューリングされるか確認",
        ],
    }

    def generate_test_view(self) -> str:
        """
        Test View（テスト観点）を生成
        
        各設定項目の固有テスト観点を提示
        """
        lines = [
            "# Test View（テスト観点）",
            "",
            f"**プロジェクト**: {self.project.name}",
            f"**生成日時**: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC",
            "",
            "---",
            "",
            "## テスト観点一覧",
            ""
        ]
        
        tested_count = 0
        total_test_cases = 0
        
        for item in self.backlog_items:
            config_item = self.config_items.get(item.config_item_id)
            if not config_item:
                continue
            
            lines.append(f"### {config_item.id}: {config_item.title}")
            lines.append("")
            
            if item.answered:
                lines.append("**ステータス**: ✅ 設定済み")
                lines.append("")
                
                lines.append("**テストケース**:")
                lines.append("")
                
                case_num = 1
                
                # 設定値の反映確認テスト
                answers = self.answers_by_config.get(config_item.id, [])
                for ans in answers:
                    val = ans.value
                    if isinstance(val, list):
                        val = ", ".join(str(v) for v in val)
                    lines.append(f"{case_num}. **{ans.input_name}** の設定値 `{val}` が正しく反映されているか確認")
                    case_num += 1
                
                # 項目固有のテスト観点を追加
                specific_tests = self._TEST_PERSPECTIVES.get(config_item.id, [])
                for test in specific_tests:
                    lines.append(f"{case_num}. {test}")
                    case_num += 1
                
                # 固有テストが無い場合はフォールバック
                if not specific_tests:
                    if config_item.description:
                        lines.append(f"{case_num}. {config_item.description}に基づく動作確認")
                        case_num += 1
                    lines.append(f"{case_num}. 関連する画面/機能での動作確認")
                    case_num += 1
                
                lines.append("")
                total_test_cases += case_num - 1
                tested_count += 1
            else:
                lines.append("**ステータス**: ⚠️ 未決定（TBD）")
                lines.append("")
                lines.append("設定が完了後、テスト観点を生成します。")
                lines.append("")
            
            lines.append("---")
            lines.append("")
        
        # サマリーを先頭に追加
        summary = [
            "## サマリー",
            "",
            f"- **テスト対象項目数**: {tested_count}/{len(self.backlog_items)}",
            f"- **テストケース総数**: {total_test_cases}",
            f"- **未決定項目数**: {len(self.backlog_items) - tested_count}",
            "",
            "---",
            ""
        ]
        
        lines = lines[:7] + summary + lines[7:]
        
        return "\n".join(lines)
    
    def generate_migration_view(self) -> str:
        """
        Migration View（移行観点）を生成
        
        移行が必要なオブジェクトを一覧化
        """
        lines = [
            "# Migration View（移行観点）",
            "",
            f"**プロジェクト**: {self.project.name}",
            f"**生成日時**: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC",
            "",
            "---",
            "",
            "## 移行対象オブジェクト",
            ""
        ]
        
        migration_items = []
        
        for item in self.backlog_items:
            config_item = self.config_items.get(item.config_item_id)
            if not config_item:
                continue
            
            # MIGRATION_VIEWを生成する設定項目のみ
            produces = config_item.produces or []
            if 'MIGRATION_VIEW' not in produces:
                continue
            
            migration_items.append((item, config_item))
        
        if not migration_items:
            lines.append("移行対象のマスタデータはありません。")
        else:
            lines.append("| 設定項目 | 移行オブジェクト | ステータス | 備考 |")
            lines.append("|---|---|---|---|")
            
            for backlog_item, config_item in migration_items:
                status = "✅ 設定済み" if backlog_item.answered else "⚠️ TBD（未決定）"
                
                # 移行オブジェクト名を推定
                migration_object = self._estimate_migration_object(config_item)
                
                # 備考
                notes = []
                if backlog_item.answered:
                    answers = self.answers_by_config.get(config_item.id, [])
                    for ans in answers:
                        notes.append(f"{ans.input_name}={ans.value}")
                note_str = "; ".join(notes) if notes else "-"
                
                lines.append(
                    f"| {config_item.title} | {migration_object} | {status} | {note_str} |"
                )
        
        lines.append("")
        lines.append("---")
        lines.append("")
        lines.append("## 移行手順")
        lines.append("")
        lines.append("1. マスタデータの抽出（旧システム）")
        lines.append("2. データクレンジング・変換")
        lines.append("3. テストデータ投入")
        lines.append("4. 整合性確認")
        lines.append("5. 本番データ移行")
        lines.append("")
        
        return "\n".join(lines)
    
    def _estimate_migration_object(self, config_item: models.ConfigItem) -> str:
        """移行オブジェクト名を推定"""
        title_lower = config_item.title.lower()
        
        if '会社' in config_item.title:
            return '会社コードマスタ'
        elif '勘定科目' in config_item.title:
            return '勘定科目マスタ'
        elif 'bp' in title_lower or '得意先' in config_item.title or '仕入先' in config_item.title:
            return 'ビジネスパートナーマスタ'
        elif '年度' in config_item.title:
            return '会計年度設定'
        else:
            return config_item.title
    
    def generate_json_export(self) -> str:
        """
        JSON形式で全決定事項・設定・回答をエクスポート
        
        LedgerForge連携を想定した構造化データ出力。
        """
        export_data = {
            "project": {
                "id": self.project.id,
                "name": self.project.name,
                "mode": self.project.mode.value if self.project.mode else "EXPERT",
                "country": self.project.country,
                "currency": self.project.currency,
                "industry": self.project.industry,
                "company_count": self.project.company_count,
                "created_at": self.project.created_at.isoformat() if self.project.created_at else None,
            },
            "exported_at": datetime.utcnow().isoformat(),
            "decisions": [],
            "config_items": [],
            "summary": {
                "total_items": len(self.backlog_items),
                "answered": len(self.answered_config_ids),
                "tbd": len(self.backlog_items) - len(self.answered_config_ids),
            }
        }
        
        # 決定事項
        for decision in self.decisions:
            config_item = self.config_items.get(decision.config_item_id)
            export_data["decisions"].append({
                "config_item_id": decision.config_item_id,
                "title": decision.title,
                "rationale": decision.rationale,
                "impact": decision.impact,
                "status": decision.status,
                "decided_at": decision.created_at.isoformat() if decision.created_at else None,
                "priority": config_item.priority if config_item else None,
            })
        
        # 設定項目と回答
        for item in self.backlog_items:
            config_item = self.config_items.get(item.config_item_id)
            if not config_item:
                continue
            
            item_data = {
                "id": config_item.id,
                "title": config_item.title,
                "priority": config_item.priority,
                "status": item.status.value,
                "answered": item.answered,
                "depends_on": config_item.depends_on or [],
                "produces": config_item.produces or [],
                "answers": {},
            }
            
            if item.answered:
                answers = self.answers_by_config.get(config_item.id, [])
                for ans in answers:
                    item_data["answers"][ans.input_name] = ans.value
            
            export_data["config_items"].append(item_data)
        
        return json.dumps(export_data, ensure_ascii=False, indent=2)

    def generate_xlsx_export(self) -> bytes:
        """
        XLSX形式で設定ワークブックをエクスポート
        
        複数シートに分けて出力:
        - サマリー: プロジェクト概要と進捗
        - 設定項目一覧: 全設定項目の詳細
        - 決定事項ログ: 全決定事項
        - テスト観点: テストケース一覧
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        
        # スタイル定義
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        def style_header_row(ws, num_cols):
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
        
        # === シート1: サマリー ===
        ws_summary = wb.active
        ws_summary.title = "サマリー"
        
        total = len(self.backlog_items)
        done = sum(1 for item in self.backlog_items if item.status == models.BacklogStatus.DONE)
        ready = sum(1 for item in self.backlog_items if item.status == models.BacklogStatus.READY)
        blocked = sum(1 for item in self.backlog_items if item.status == models.BacklogStatus.BLOCKED)
        
        summary_data = [
            ("プロジェクト名", self.project.name),
            ("モード", self.project.mode.value if self.project.mode else "EXPERT"),
            ("国", self.project.country or "-"),
            ("通貨", self.project.currency or "-"),
            ("業種", self.project.industry or "-"),
            ("会社数", self.project.company_count or "-"),
            ("", ""),
            ("全設定項目数", total),
            ("完了", done),
            ("対応可能", ready),
            ("ブロック中", blocked),
            ("完了率", f"{round(done / total * 100, 1)}%" if total > 0 else "0%"),
            ("", ""),
            ("生成日時", datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')),
        ]
        
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 40
        
        for row_idx, (label, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True) if label else Font()
            ws_summary.cell(row=row_idx, column=2, value=value)
        
        # === シート2: 設定項目一覧 ===
        ws_config = wb.create_sheet("設定項目一覧")
        headers = ["ID", "タイトル", "優先度", "ステータス", "回答済み", "依存関係", "設定値"]
        for col_idx, header in enumerate(headers, 1):
            ws_config.cell(row=1, column=col_idx, value=header)
        style_header_row(ws_config, len(headers))
        
        row_idx = 2
        for item in self.backlog_items:
            config_item = self.config_items.get(item.config_item_id)
            if not config_item:
                continue
            
            depends_str = ", ".join(config_item.depends_on or []) or "-"
            
            if item.answered:
                answers = self.answers_by_config.get(config_item.id, [])
                value_parts = []
                for ans in answers:
                    val = ans.value
                    if isinstance(val, list):
                        val = ", ".join(str(v) for v in val)
                    value_parts.append(f"{ans.input_name}={val}")
                value_str = "; ".join(value_parts) if value_parts else "設定済み"
            else:
                value_str = "TBD（未決定）"
            
            ws_config.cell(row=row_idx, column=1, value=config_item.id)
            ws_config.cell(row=row_idx, column=2, value=config_item.title)
            ws_config.cell(row=row_idx, column=3, value=config_item.priority)
            ws_config.cell(row=row_idx, column=4, value=item.status.value)
            ws_config.cell(row=row_idx, column=5, value="済" if item.answered else "未")
            ws_config.cell(row=row_idx, column=6, value=depends_str)
            ws_config.cell(row=row_idx, column=7, value=value_str)
            row_idx += 1
        
        # 列幅調整
        col_widths = [15, 30, 10, 12, 10, 25, 40]
        for i, width in enumerate(col_widths, 1):
            ws_config.column_dimensions[chr(64 + i)].width = width
        
        # === シート3: 決定事項ログ ===
        ws_decisions = wb.create_sheet("決定事項ログ")
        decision_headers = ["No.", "設定項目ID", "タイトル", "決定内容", "影響範囲", "ステータス", "決定日時"]
        for col_idx, header in enumerate(decision_headers, 1):
            ws_decisions.cell(row=1, column=col_idx, value=header)
        style_header_row(ws_decisions, len(decision_headers))
        
        for i, decision in enumerate(self.decisions, 1):
            ws_decisions.cell(row=i + 1, column=1, value=i)
            ws_decisions.cell(row=i + 1, column=2, value=decision.config_item_id)
            ws_decisions.cell(row=i + 1, column=3, value=decision.title)
            ws_decisions.cell(row=i + 1, column=4, value=decision.rationale or "-")
            ws_decisions.cell(row=i + 1, column=5, value=decision.impact or "-")
            ws_decisions.cell(row=i + 1, column=6, value=decision.status)
            ws_decisions.cell(row=i + 1, column=7, value=decision.created_at.strftime('%Y-%m-%d %H:%M') if decision.created_at else "-")
        
        decision_col_widths = [6, 15, 25, 40, 30, 12, 18]
        for i, width in enumerate(decision_col_widths, 1):
            ws_decisions.column_dimensions[chr(64 + i)].width = width
        
        # === シート4: テスト観点 ===
        ws_test = wb.create_sheet("テスト観点")
        test_headers = ["設定項目ID", "タイトル", "ステータス", "テスト観点", "入力名", "設定値"]
        for col_idx, header in enumerate(test_headers, 1):
            ws_test.cell(row=1, column=col_idx, value=header)
        style_header_row(ws_test, len(test_headers))
        
        test_row = 2
        for item in self.backlog_items:
            config_item = self.config_items.get(item.config_item_id)
            if not config_item:
                continue
            
            if item.answered:
                answers = self.answers_by_config.get(config_item.id, [])
                for ans in answers:
                    ws_test.cell(row=test_row, column=1, value=config_item.id)
                    ws_test.cell(row=test_row, column=2, value=config_item.title)
                    ws_test.cell(row=test_row, column=3, value="設定済み")
                    ws_test.cell(row=test_row, column=4, value=f"{ans.input_name}の設定値が正しく反映されているか確認")
                    ws_test.cell(row=test_row, column=5, value=ans.input_name)
                    val = ans.value
                    if isinstance(val, list):
                        val = ", ".join(str(v) for v in val)
                    ws_test.cell(row=test_row, column=6, value=str(val))
                    test_row += 1
            else:
                ws_test.cell(row=test_row, column=1, value=config_item.id)
                ws_test.cell(row=test_row, column=2, value=config_item.title)
                ws_test.cell(row=test_row, column=3, value="TBD（未決定）")
                ws_test.cell(row=test_row, column=4, value="設定完了後にテスト観点を生成")
                test_row += 1
        
        test_col_widths = [15, 30, 12, 50, 20, 25]
        for i, width in enumerate(test_col_widths, 1):
            ws_test.column_dimensions[chr(64 + i)].width = width
        
        # バイト列として返す
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def generate_all(self) -> Dict[models.ArtifactType, tuple[str, int]]:
        """
        全ての成果物を生成
        
        Returns:
            {ArtifactType: (content, tbd_count)} の辞書
        """
        artifacts = {}
        
        # Decision Log
        decision_log = self.generate_decision_log()
        artifacts[models.ArtifactType.DECISION_LOG] = (
            decision_log,
            self._count_tbd(decision_log)
        )
        
        # Config Workbook
        config_workbook = self.generate_config_workbook()
        artifacts[models.ArtifactType.CONFIG_WORKBOOK] = (
            config_workbook,
            self._count_tbd(config_workbook)
        )
        
        # Test View
        test_view = self.generate_test_view()
        artifacts[models.ArtifactType.TEST_VIEW] = (
            test_view,
            self._count_tbd(test_view)
        )
        
        # Migration View
        migration_view = self.generate_migration_view()
        artifacts[models.ArtifactType.MIGRATION_VIEW] = (
            migration_view,
            self._count_tbd(migration_view)
        )
        
        return artifacts


def generate_artifacts(db: Session, project_id: int, artifact_types: List[models.ArtifactType] = None) -> List[models.Artifact]:
    """
    成果物を生成してDBに保存
    
    Args:
        db: データベースセッション
        project_id: プロジェクトID
        artifact_types: 生成する成果物の種類（Noneの場合は全て）
        
    Returns:
        生成された成果物のリスト
    """
    generator = ArtifactGenerator(db, project_id)
    all_artifacts = generator.generate_all()
    
    # 指定された種類のみフィルタ
    if artifact_types:
        all_artifacts = {k: v for k, v in all_artifacts.items() if k in artifact_types}
    
    # DBに保存
    saved_artifacts = []
    for artifact_type, (content, tbd_count) in all_artifacts.items():
        artifact = crud.create_artifact(db, project_id, artifact_type, content, tbd_count)
        saved_artifacts.append(artifact)
    
    return saved_artifacts
